import openpyxl

# -------     Write Data from excel   ---------
def write_data():
    file = 'C:/Users/MOHIT/OneDrive/Desktop/simpla_data_read.xlsx'
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for r in range(1,6):
        for c in range(1,5):
            sheet.cell(r,c).value="Welcome to Bridgelabz"

    workbook.save(file)

def write_data_in_each_cell():
    file = 'C:/Users/MOHIT/OneDrive/Desktop/simpla_data_read.xlsx'
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["mohit1"]

    sheet.cell(1,1).value=101
    sheet.cell(1,2).value=102
    sheet.cell(1,3).value=103
    sheet.cell(1,4).value=104
    sheet.cell(1,5).value=105
    sheet.cell(1,6).value=106

    sheet.cell(2, 1).value = "Mohit"
    sheet.cell(2, 2).value = "Swapnil"
    sheet.cell(2, 3).value = "Ganesh"
    sheet.cell(2, 4).value = "Monu"
    sheet.cell(2, 5).value = "sonu"
    sheet.cell(2, 6).value = "mani"

    sheet.cell(3, 1).value = "Engineer"
    sheet.cell(3, 2).value = "Doctor"
    sheet.cell(3, 3).value = "Lawyer"
    sheet.cell(3, 4).value = "Builder"
    sheet.cell(3, 5).value = "Mechnical"
    sheet.cell(3, 6).value = "Gardener"

    workbook.save(file)

# write_data()
write_data_in_each_cell()