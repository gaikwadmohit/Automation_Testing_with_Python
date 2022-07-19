import openpyxl

# -------     Read Data from excel   ---------
file = 'C:/Users/MOHIT/OneDrive/Desktop/simpla_data_read.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook["mohit"]

rows = sheet.max_row
cols = sheet.max_column

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end='                  || ')
    print()
